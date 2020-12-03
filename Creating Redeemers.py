# -*- coding: utf-8 -*-
"""
Created on Tue Jan 21 11:30:00 2020

@author: matth
"""

import xlrd


file = ('Qlik Table Data.xlsx')
wb = xlrd.open_workbook(file)
sheet = wb.sheet_by_index(2)

accounts_demo = {}

for i in range(1, sheet.nrows):
    accounts_demo[str(sheet.cell_value(i, 3))] = "N"


sheet = wb.sheet_by_index(3)

accounts_redem = []
for i in range(1, sheet.nrows):
    accounts_redem.append(str(sheet.cell_value(i, 2)))


for acct in accounts_demo:
    if acct in accounts_redem:
        accounts_demo[acct] = "Y"





import openpyxl

# Create the workbook and sheet for Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# openpyxl does things based on 1 instead of 0
row = 2
for key,values in accounts_demo.items():
    # Put the key in the first column for each key in the dictionary
    sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values:
        # Put the element in each adjacent column for each element in the tuple
        sheet.cell(row=row, column=column, value=element)
        column += 1
    row += 1

workbook.save(filename="my_workbook.xlsx")


    
